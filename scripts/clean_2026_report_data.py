#!/usr/bin/env python3
import argparse
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path()
OUTPUT_DIR = Path()
CLEANED_TRACKER = OUTPUT_DIR / "report_data_cleaned.csv"
SUMMARY = OUTPUT_DIR / "cleanup_summary.md"
OUTCOME_SUMMARY = OUTPUT_DIR / "dashboard_summary_cleaned.csv"
REVIEW_QUEUE = OUTPUT_DIR / "review_queue.csv"
NORMALIZATION_AUDIT = OUTPUT_DIR / "normalization_audit.csv"
MANIFEST = OUTPUT_DIR / "release_manifest.json"

IFERROR_STRING_FALLBACK = re.compile(r'\),"((?:""|[^"])*)"\)$')
IFERROR_NUMBER_FALLBACK = re.compile(r"\),([0-9]+(?:\.[0-9]+)?)\)$")

EMPLOYER_ALIASES = {
    "amazon inc": "Amazon",
    "amazon inc.": "Amazon",
    "fidelity": "Fidelity Investments",
    "fidelity investment": "Fidelity Investments",
    "google llc": "Google",
    "meta platforms": "Meta",
    "meta platforms inc": "Meta",
    "oracle corporation": "Oracle",
    "capital one financial": "Capital One",
    "capital one financial corporation": "Capital One",
    "liberty mutual insurance": "Liberty Mutual",
    "the travelers companies": "Travelers",
    "the travelers companies, inc.": "Travelers",
    "nvidia corporation": "NVIDIA",
    "international business machines": "IBM",
    "apple inc": "Apple",
    "apple inc.": "Apple",
    "mathworks": "MathWorks",
    "the mathworks": "MathWorks",
    "the mathworks, inc.": "MathWorks",
    "athenahealth": "athenahealth",
    "athena health": "athenahealth",
    "guideware software": "Guidewire",
}

SCHOOL_ALIASES = {
    "cics": "UMass Amherst",
    "manning cics": "UMass Amherst",
    "manning cics, umass": "UMass Amherst",
    "manning cics, umass amherst": "UMass Amherst",
    "umass": "UMass Amherst",
    "umass, amherst": "UMass Amherst",
    "umass amherst": "UMass Amherst",
    "umass cics": "UMass Amherst",
    "umass graduate school": "UMass Amherst",
    "umass amherst isenberg msba": "UMass Amherst",
    "university of massachusetts amherst": "UMass Amherst",
    "northeastern ms cs": "Northeastern University",
    "georgia institute of technology": "Georgia Institute of Technology",
    "the university of texas at dallas": "University of Texas at Dallas",
    "manning college of information and computer sciences, umass": "UMass Amherst",
    "manning college of information and computer sciences, umass amherst": "UMass Amherst",
    "columbia": "Columbia University",
    "uiuc": "University of Illinois Urbana-Champaign",
    "nyu courant (cds)": "NYU",
    "carnegie mellon": "Carnegie Mellon University",
}

MAJOR_ALIASES = {
    "computer science(bs)": "Computer Science BS",
    "computer science(ba)": "Computer Science BA",
    "informatics(bs)": "Informatics BS",
}

ROLE_ALIASES = {
    "software engineering intern": "Software Engineer Intern",
    "software development engineer intern": "Software Engineer Intern",
    "sde intern": "Software Engineer Intern",
    "software engineer internship": "Software Engineer Intern",
    "software development engineer": "Software Development Engineer",
    "sde": "Software Development Engineer",
    "ml engineer": "ML Engineer",
    "machine learning engineer": "Machine Learning Engineer",
    "ai engineer": "AI Engineer",
    "ai software engineer": "AI Software Engineer",
    "data scientist": "Data Scientist",
    "data engineer": "Data Engineer",
    "data analyst": "Data Analyst",
    "research scientist": "Research Scientist",
    "research intern": "Research Intern",
    "applied scientist": "Applied Scientist",
    "applied scientist ii": "Applied Scientist II",
    "product manager": "Product Manager",
    "customer success sales engineer": "Customer Success Sales Engineer",
    "data analyst? liz is filling this out": "Data Analyst",
    "swe": "Software Engineer",
}

STATE_ALIASES = {
    "A": "MA",
    "AD": "MD",
    "AS": "TX",
    "HO": "ID",
    "INT'L": "",
    "INTL": "",
    "LL": "MA",
    "N/A": "",
    "NA": "",
    "CALIFORNIA": "CA",
    "MASSACHUSETTS": "MA",
    "WASHINGTON": "WA",
    "NEW HAMPSHIRE": "NH",
    "VIRGINIA": "VA",
    "COLORADO": "CO",
    "MARYLAND": "MD",
    "MAINE": "ME",
    "MINNESOTA": "MN",
    "TEXAS": "TX",
    "CONNECTICUT": "CT",
    "NEW YORK": "NY",
    "RHODE ISLAND": "RI",
}

STATE_FROM_LOCATION = re.compile(r"(?:,\s*|\s+)([A-Z]{2})\s*$")

REQUIRED_TRACKER_HEADERS = {
    "Program",
    "Email",
    "Degree Confer Date",
    "Outcomes Status",
    "Employer",
    "Role",
    "Location",
    "Location - State",
    "School Name",
    "Degree",
    "Majors Name",
    "LinkedIn",
    "LI LookUp Date",
    "Careers Survey Timestamp",
    "Interactions",
    "Doc Reviews",
}


def normalize_space(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def formula_fallback(value):
    if value is None:
        return ""
    if not isinstance(value, str) or not value.startswith("=IFERROR("):
        return normalize_space(value)

    match = IFERROR_STRING_FALLBACK.search(value)
    if match:
        return normalize_space(match.group(1).replace('""', '"'))

    match = IFERROR_NUMBER_FALLBACK.search(value)
    if match:
        return normalize_space(match.group(1))

    return ""


def is_missing(value):
    return normalize_space(value).lower() in {"", "#n/a", "unknown", "none found", "n/a"}


def title_keep_acronyms(value):
    words = []
    for word in normalize_space(value).split(" "):
        if word.upper() in {"AI", "ML", "SDE", "QA", "EDG", "MTS"}:
            words.append(word.upper())
        elif word.upper() in {"NVIDIA", "IBM", "NASA", "PTC", "UKG"}:
            words.append(word.upper())
        else:
            words.append(word[:1].upper() + word[1:])
    return " ".join(words)


def canonicalize(value, aliases):
    text = normalize_space(value)
    if is_missing(text):
        return ""
    key = text.lower().strip(" .")
    if key in aliases:
        return aliases[key]
    return title_keep_acronyms(text)


def clean_outcome(value):
    text = normalize_space(value)
    if text == "Searching":
        return "Looking"
    if text == "Continuing Ed":
        return "Continuing Education"
    if text == "Not seeking":
        return "Not Seeking"
    return text


def clean_degree(value):
    text = normalize_space(value)
    if is_missing(text):
        return ""
    if text == "MSMS":
        return "MS"
    if text.startswith("Other"):
        return normalize_space(text.replace("Other", "", 1))
    return text


def clean_state(state_value, location_value):
    state = normalize_space(state_value).upper()
    if state.startswith("="):
        state = ""
    location = normalize_space(location_value)
    if state == "IA" and "india" in location.lower():
        return ""
    state = STATE_ALIASES.get(state, state)
    if state:
        return state

    match = STATE_FROM_LOCATION.search(location)
    if match:
        return match.group(1).upper()
    return ""


def read_tracker_rows():
    formula_wb = load_workbook(SOURCE, read_only=True, data_only=False)
    value_wb = load_workbook(SOURCE, read_only=True, data_only=True)

    formula_ws = formula_wb["TRACKER"]
    value_ws = value_wb["TRACKER"]

    formula_iter = formula_ws.iter_rows(values_only=True)
    value_iter = value_ws.iter_rows(values_only=True)
    headers = list(next(formula_iter))
    next(value_iter)

    for row_number, (formula_row, value_row) in enumerate(zip(formula_iter, value_iter), start=2):
        if not any(value is not None for value in formula_row):
            continue
        formula_record = dict(zip(headers, formula_row))
        value_record = dict(zip(headers, value_row))

        cached = {header: normalize_space(value_record.get(header)) for header in headers}
        raw = {}
        for header in headers:
            formula_value = formula_record.get(header)
            value = formula_fallback(formula_value)
            # Array formulas are returned as openpyxl objects rather than strings.
            # Their cached values are the reportable values in this workbook.
            if cached.get(header) and (
                not isinstance(formula_value, (str, int, float, bool, datetime))
                or (isinstance(formula_value, str) and formula_value.startswith("="))
            ):
                value = cached[header]
            raw[header] = value
        yield row_number, raw, cached


def validate_input_workbook():
    """Fail early when the approved workbook no longer matches the data contract."""
    workbook = load_workbook(SOURCE, read_only=True, data_only=False)
    if "TRACKER" not in workbook.sheetnames:
        raise ValueError("Input workbook is missing the required TRACKER sheet")

    worksheet = workbook["TRACKER"]
    headers = {normalize_space(value) for value in next(worksheet.iter_rows(values_only=True))}
    missing_headers = sorted(REQUIRED_TRACKER_HEADERS - headers)
    if missing_headers:
        raise ValueError(
            "Input workbook TRACKER sheet is missing required columns: "
            + ", ".join(missing_headers)
        )


def sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_record_key(email):
    """Private, deterministic key for matching protected exports across releases."""
    return hashlib.sha256(normalize_space(email).lower().encode("utf-8")).hexdigest()


def read_dashboard_work_auth_rows():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    if "Dashboard_Data" not in workbook.sheetnames:
        return []
    sheet = workbook["Dashboard_Data"]
    rows = list(sheet.iter_rows(values_only=True))
    start = None
    for index, row in enumerate(rows):
        if row and row[0] == "Work Authorization Group":
            start = index + 1
            break

    if start is None:
        return []

    summary_rows = []
    for row in rows[start:]:
        if not row or not row[0]:
            break
        summary_rows.append({
            "group": row[0],
            "total": row[1] or 0,
            "Working": row[2] or 0,
            "Looking": row[3] or 0,
            "Continuing Education": row[4] or 0,
            "Unknown": row[5] or 0,
            "Not Seeking": row[6] or 0,
        })
    return summary_rows


def build_outputs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    validate_input_workbook()

    cleaned_rows = []
    for row_number, raw, cached in read_tracker_rows():
        outcome_report = clean_outcome(raw["Outcomes Status"])
        employer_report = canonicalize(raw["Employer"], EMPLOYER_ALIASES)
        role_report = canonicalize(raw["Role"], ROLE_ALIASES)
        school_report = canonicalize(raw["School Name"], SCHOOL_ALIASES)
        degree_report = clean_degree(raw["Degree"])
        major_report = canonicalize(raw["Majors Name"], MAJOR_ALIASES)
        state_report = clean_state(raw["Location - State"], raw["Location"])
        flags = []
        if outcome_report == "Working" and not employer_report:
            flags.append("working_missing_employer")
        if outcome_report == "Working" and not role_report:
            flags.append("working_missing_role")
        if outcome_report == "Continuing Education" and not school_report:
            flags.append("continuing_ed_missing_school")
        if outcome_report == "Working" and not state_report:
            flags.append("working_missing_state")

        cleaned_rows.append({
            "source_row": row_number,
            "record_key": stable_record_key(raw["Email"]),
            "program": raw["Program"],
            "degree_confer_date": raw["Degree Confer Date"],
            "outcome_raw": raw["Outcomes Status"],
            "outcome_report": outcome_report,
            "known_outcome": outcome_report not in {"Unknown"},
            "positive_outcome": outcome_report in {"Working", "Continuing Education"},
            "employer_raw": raw["Employer"],
            "employer_report": employer_report,
            "role_raw": raw["Role"],
            "role_report": role_report,
            "school_raw": raw["School Name"],
            "school_report": school_report,
            "degree_raw": raw["Degree"],
            "degree_report": degree_report,
            "major_raw": raw["Majors Name"],
            "major_report": major_report,
            "location_raw": raw["Location"],
            "state_raw": raw["Location - State"],
            "state_report": state_report,
            "has_linkedin": bool(raw["LinkedIn"]) and raw["LinkedIn"].lower() not in {"none found"},
            "has_li_lookup_date": bool(raw["LI LookUp Date"]),
            "has_survey_timestamp": bool(raw["Careers Survey Timestamp"]) and raw["Careers Survey Timestamp"] != "#N/A",
            "has_interaction": bool(raw["Interactions"]) and raw["Interactions"] != "#N/A",
            "has_doc_review_signal": bool(raw["Doc Reviews"]) and raw["Doc Reviews"] != "#N/A",
            "cleanup_flags": ";".join(flags),
        })

    fieldnames = list(cleaned_rows[0].keys())
    with CLEANED_TRACKER.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(cleaned_rows)

    write_dashboard_summary(cleaned_rows)
    write_summary(cleaned_rows)
    write_review_queue(cleaned_rows)
    write_normalization_audit(cleaned_rows)
    write_manifest(cleaned_rows)


def counter_for(rows, field):
    return Counter(row[field] for row in rows if row[field])


def write_dashboard_summary(rows):
    with OUTCOME_SUMMARY.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["section", "group", "total", "Working", "Looking", "Continuing Education", "Unknown", "Not Seeking"])

        def write_group(section, group, group_rows):
            counts = Counter(row["outcome_report"] for row in group_rows)
            writer.writerow([
                section,
                group,
                len(group_rows),
                counts["Working"],
                counts["Looking"],
                counts["Continuing Education"],
                counts["Unknown"],
                counts["Not Seeking"],
            ])

        write_group("all", "All Students", rows)
        for program in ["Undergraduate", "MS", "PhD"]:
            write_group("program", program, [row for row in rows if row["program"] == program])
        for date in sorted(counter_for(rows, "degree_confer_date")):
            write_group("degree_confer_date", date, [row for row in rows if row["degree_confer_date"] == date])
        for row in read_dashboard_work_auth_rows():
            writer.writerow([
                "work_authorization",
                row["group"],
                row["total"],
                row["Working"],
                row["Looking"],
                row["Continuing Education"],
                row["Unknown"],
                row["Not Seeking"],
            ])


def write_review_queue(rows):
    """Write only records needing human review to the private output directory."""
    fields = [
        "source_row", "program", "degree_confer_date", "outcome_report",
        "employer_raw", "employer_report", "role_raw", "role_report",
        "school_raw", "school_report", "location_raw", "state_raw",
        "state_report", "major_raw", "major_report", "cleanup_flags",
    ]
    with REVIEW_QUEUE.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(
            {field: row[field] for field in fields}
            for row in rows
            if row["cleanup_flags"]
        )


def write_normalization_audit(rows):
    """Summarize every changed value so reviewers can approve the cleaning rules."""
    mappings = [
        ("outcome", "outcome_raw", "outcome_report"),
        ("employer", "employer_raw", "employer_report"),
        ("role", "role_raw", "role_report"),
        ("school", "school_raw", "school_report"),
        ("degree", "degree_raw", "degree_report"),
        ("major", "major_raw", "major_report"),
        ("state", "state_raw", "state_report"),
    ]
    counts = Counter()
    for row in rows:
        for field, raw_key, normalized_key in mappings:
            raw_value = row[raw_key]
            normalized_value = row[normalized_key]
            if raw_value and raw_value != normalized_value:
                counts[(field, raw_value, normalized_value)] += 1

    with NORMALIZATION_AUDIT.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["field", "raw_value", "normalized_value", "record_count"],
        )
        writer.writeheader()
        for (field, raw_value, normalized_value), count in sorted(counts.items()):
            writer.writerow({
                "field": field,
                "raw_value": raw_value,
                "normalized_value": normalized_value,
                "record_count": count,
            })


def write_manifest(rows):
    manifest = {
        "source_file": SOURCE.name,
        "source_sha256": sha256(SOURCE),
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "row_count": len(rows),
        "outputs": [
            CLEANED_TRACKER.name,
            OUTCOME_SUMMARY.name,
            SUMMARY.name,
            REVIEW_QUEUE.name,
            NORMALIZATION_AUDIT.name,
        ],
    }
    MANIFEST.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

def markdown_count_table(title, counts, limit=15):
    lines = [f"### {title}", "", "| Value | Count |", "|---|---:|"]
    for value, count in counts.most_common(limit):
        lines.append(f"| {value} | {count} |")
    lines.append("")
    return lines


def write_summary(rows):
    working_rows = [row for row in rows if row["outcome_report"] == "Working"]
    continuing_rows = [row for row in rows if row["outcome_report"] == "Continuing Education"]

    flag_counts = Counter()
    for row in rows:
        for flag in row["cleanup_flags"].split(";"):
            if flag:
                flag_counts[flag] += 1

    lines = [
        "# 2026 Report Data Cleanup Summary",
        "",
        f"Source workbook: `{SOURCE}`",
        f"Cleaned row-level export: `{CLEANED_TRACKER}`",
        f"Dashboard summary export: `{OUTCOME_SUMMARY}`",
        "",
        "## Coverage",
        "",
        f"- Total rows: {len(rows)}",
        f"- Working rows with report-ready employer: {sum(1 for row in working_rows if row['employer_report'])} / {len(working_rows)}",
        f"- Working rows with report-ready role: {sum(1 for row in working_rows if row['role_report'])} / {len(working_rows)}",
        f"- Working rows with report-ready state: {sum(1 for row in working_rows if row['state_report'])} / {len(working_rows)}",
        f"- Continuing education rows with report-ready school: {sum(1 for row in continuing_rows if row['school_report'])} / {len(continuing_rows)}",
        "",
        "## Cleanup Flags",
        "",
        "| Flag | Count |",
        "|---|---:|",
    ]
    for flag, count in flag_counts.most_common():
        lines.append(f"| {flag} | {count} |")
    lines.append("")

    lines += markdown_count_table("Outcome Names", counter_for(rows, "outcome_report"))
    lines += markdown_count_table("Top Employers", counter_for(working_rows, "employer_report"), 20)
    lines += markdown_count_table("Top Roles", counter_for(working_rows, "role_report"), 20)
    lines += markdown_count_table("Continuing Education Schools", counter_for(continuing_rows, "school_report"), 20)
    lines += markdown_count_table("Working States", counter_for(working_rows, "state_report"), 20)

    SUMMARY.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Clean a sensitive CICS outcomes workbook into private review artifacts."
    )
    parser.add_argument("--input", type=Path, required=True, help="Path to the approved source workbook")
    parser.add_argument(
        "--output-dir",
        type=Path,
        required=True,
        help="Private directory for generated row-level and review artifacts",
    )
    args = parser.parse_args()

    if not args.input.is_file():
        parser.error(f"input workbook not found: {args.input}")

    SOURCE = args.input.resolve()
    OUTPUT_DIR = args.output_dir.resolve()
    CLEANED_TRACKER = OUTPUT_DIR / "report_data_cleaned.csv"
    SUMMARY = OUTPUT_DIR / "cleanup_summary.md"
    OUTCOME_SUMMARY = OUTPUT_DIR / "dashboard_summary_cleaned.csv"
    REVIEW_QUEUE = OUTPUT_DIR / "review_queue.csv"
    NORMALIZATION_AUDIT = OUTPUT_DIR / "normalization_audit.csv"
    MANIFEST = OUTPUT_DIR / "release_manifest.json"
    build_outputs()
