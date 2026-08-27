#!/usr/bin/env python3
"""Create private, non-identifying employer aggregates from Employer Dashboard data."""

import argparse
import csv
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROGRAMS = {
    "Bachelors": "Undergraduate",
    "Masters": "MS",
    "Doctorate": "PhD",
}
REQUIRED_HEADERS = {"FDS Year", "Outcome", "Education Level", "Employer"}


def normalize(value):
    return str(value or "").strip()


def year_matches(value, year):
    text = normalize(value)
    return text == str(year) or text == f"{year}.0"


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Approved Employer Dashboard workbook")
    parser.add_argument("--year", type=int, required=True, help="FDS year to review")
    parser.add_argument("--summary-output", type=Path, required=True, help="Private program summary CSV")
    parser.add_argument("--employers-output", type=Path, required=True, help="Private employer count CSV")
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    missing_sheets = {"Data", "Employers"} - set(workbook.sheetnames)
    if missing_sheets:
        raise ValueError("Workbook is missing required sheets: " + ", ".join(sorted(missing_sheets)))

    register_sheet = workbook["Employers"]
    curated_employers = {
        normalize(row[0]).casefold()
        for row in register_sheet.iter_rows(min_row=2, values_only=True)
        if row and normalize(row[0])
    }

    data_sheet = workbook["Data"]
    headers = list(next(data_sheet.iter_rows(values_only=True)))
    positions = {normalize(header): index for index, header in enumerate(headers)}
    missing_headers = REQUIRED_HEADERS - positions.keys()
    if missing_headers:
        raise ValueError("Data is missing required columns: " + ", ".join(sorted(missing_headers)))

    employers = defaultdict(Counter)
    missing_from_register = Counter()
    unmapped_education_levels = Counter()
    for row in data_sheet.iter_rows(min_row=2, values_only=True):
        if not year_matches(row[positions["FDS Year"]], args.year):
            continue
        if normalize(row[positions["Outcome"]]).casefold() != "job":
            continue
        program = PROGRAMS.get(normalize(row[positions["Education Level"]]))
        if not program:
            unmapped_education_levels[normalize(row[positions["Education Level"]]) or "(blank)"] += 1
            continue
        employer = normalize(row[positions["Employer"]])
        if not employer:
            continue
        employers[program][employer] += 1
        if employer.casefold() not in curated_employers:
            missing_from_register[program] += 1

    args.summary_output.parent.mkdir(parents=True, exist_ok=True)
    args.employers_output.parent.mkdir(parents=True, exist_ok=True)
    with args.summary_output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=[
            "program", "dashboard_job_records", "unique_employers",
            "employers_hiring_multiple", "max_hires_at_one_employer",
            "employer_records_not_in_curated_register", "unmapped_education_level_records",
            "method",
        ])
        writer.writeheader()
        for program in ("Undergraduate", "MS", "PhD"):
            counts = employers[program]
            writer.writerow({
                "program": program,
                "dashboard_job_records": sum(counts.values()),
                "unique_employers": len(counts),
                "employers_hiring_multiple": sum(count > 1 for count in counts.values()),
                "max_hires_at_one_employer": max(counts.values(), default=0),
                "employer_records_not_in_curated_register": missing_from_register[program],
                "unmapped_education_level_records": sum(unmapped_education_levels.values()),
                "method": f"Employer Dashboard Data; FDS Year {args.year}; Outcome = Job",
            })

    with args.employers_output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["program", "employer", "job_records", "hired_multiple"])
        writer.writeheader()
        for program in ("Undergraduate", "MS", "PhD"):
            for employer, count in sorted(employers[program].items(), key=lambda item: (-item[1], item[0].casefold())):
                writer.writerow({
                    "program": program,
                    "employer": employer,
                    "job_records": count,
                    "hired_multiple": count > 1,
                })


if __name__ == "__main__":
    main()
