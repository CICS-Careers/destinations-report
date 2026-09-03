#!/usr/bin/env python3
"""Create a private, non-identifying salary-range review aggregate."""

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from clean_2026_report_data import formula_fallback, normalize_space  # noqa: E402


PROGRAMS = ("Undergraduate", "MS", "PhD")
MIN_ANNUAL_SALARY = 30_000
MAX_ANNUAL_SALARY = 500_000


def percentile(values, proportion):
    """Linear-interpolated percentile, matching spreadsheet percentile calculations."""
    position = (len(values) - 1) * proportion
    lower = int(position)
    upper = min(lower + 1, len(values) - 1)
    return values[lower] + (values[upper] - values[lower]) * (position - lower)


def salary_value(value):
    value = formula_fallback(value)
    try:
        return float(normalize_space(value).replace("$", "").replace(",", ""))
    except (TypeError, ValueError):
        return None


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Approved source workbook")
    parser.add_argument("--output", type=Path, required=True, help="Private aggregate CSV")
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=False)
    sheet = workbook["TRACKER"]
    headers = list(next(sheet.values))
    positions = {header: index for index, header in enumerate(headers)}
    required = {"Program", "Outcomes Status", "Career Survey Base Salary"}
    missing = required - positions.keys()
    if missing:
        raise ValueError("TRACKER is missing: " + ", ".join(sorted(missing)))

    results = []
    for program in PROGRAMS:
        working = 0
        usable = []
        implausible = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[positions["Program"]] != program or normalize_space(row[positions["Outcomes Status"]]) != "Working":
                continue
            working += 1
            value = salary_value(row[positions["Career Survey Base Salary"]])
            if value is None:
                continue
            if not MIN_ANNUAL_SALARY <= value <= MAX_ANNUAL_SALARY:
                implausible += 1
                continue
            usable.append(value)
        usable.sort()
        results.append({
            "program": program,
            "working_records": working,
            "usable_annual_base_salary_records": len(usable),
            "coverage_percent": round(100 * len(usable) / working, 1) if working else 0,
            "excluded_implausible_values": implausible,
            "p25": round(percentile(usable, 0.25), 2) if usable else "",
            "p50": round(percentile(usable, 0.50), 2) if usable else "",
            "p75": round(percentile(usable, 0.75), 2) if usable else "",
            "method": "Working records; annual base salary $30,000-$500,000; linear 25th/50th/75th percentiles",
        })

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=results[0].keys())
        writer.writeheader()
        writer.writerows(results)


if __name__ == "__main__":
    main()
