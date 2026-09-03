#!/usr/bin/env python3
"""Privately reconcile two Hypercare TRACKER sheets using stable email keys."""

import argparse
import csv
import hashlib
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


KEY_FIELD = "Email"
COMPARE_FIELDS = (
    "Program",
    "Degree Confer Date",
    "Outcomes Status",
    "Employer",
    "Role",
    "Location",
    "Location - State",
    "School Name",
    "Degree",
)


def clean(value):
    return "" if value is None else " ".join(str(value).strip().split())


def read_tracker(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "TRACKER" not in workbook.sheetnames:
        raise ValueError(f"{path} is missing the TRACKER sheet")
    rows = workbook["TRACKER"].iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    missing = {KEY_FIELD, *COMPARE_FIELDS} - set(headers)
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(sorted(missing))}")

    records = {}
    duplicates = []
    for row_number, row in enumerate(rows, start=2):
        record = dict(zip(headers, row))
        email = clean(record[KEY_FIELD]).lower()
        if not email:
            continue
        if email in records:
            duplicates.append(email)
        records[email] = {
            "source_row": row_number,
            **{field: clean(record[field]) for field in COMPARE_FIELDS},
        }
    return records, duplicates


def main():
    parser = argparse.ArgumentParser(
        description="Create private stable-key reconciliation artifacts for two Hypercare workbooks."
    )
    parser.add_argument("--previous", type=Path, required=True, help="Earlier approved workbook")
    parser.add_argument("--current", type=Path, required=True, help="Candidate workbook")
    parser.add_argument("--output-dir", type=Path, required=True, help="Private reconciliation output folder")
    args = parser.parse_args()

    if not args.previous.is_file() or not args.current.is_file():
        parser.error("both --previous and --current must be existing workbooks")

    previous, previous_duplicates = read_tracker(args.previous)
    current, current_duplicates = read_tracker(args.current)
    added = sorted(set(current) - set(previous))
    removed = sorted(set(previous) - set(current))
    common = sorted(set(previous) & set(current))

    changed = []
    changed_fields = Counter()
    for email in common:
        fields = [field for field in COMPARE_FIELDS if previous[email][field] != current[email][field]]
        if fields:
            changed.append((email, fields))
            changed_fields.update(fields)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    summary = {
        "previous_file": args.previous.name,
        "current_file": args.current.name,
        "previous_records": len(previous),
        "current_records": len(current),
        "added_records": len(added),
        "removed_records": len(removed),
        "changed_records": len(changed),
        "changed_field_counts": dict(changed_fields),
        "duplicate_email_counts": {
            "previous": len(previous_duplicates),
            "current": len(current_duplicates),
        },
    }
    (args.output_dir / "reconciliation_summary.json").write_text(
        json.dumps(summary, indent=2) + "\n", encoding="utf-8"
    )

    detail_fields = ["change_type", "email", "previous_source_row", "current_source_row", "changed_fields"]
    with (args.output_dir / "reconciliation_detail.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=detail_fields)
        writer.writeheader()
        for email in added:
            writer.writerow({"change_type": "added", "email": email, "current_source_row": current[email]["source_row"]})
        for email in removed:
            writer.writerow({"change_type": "removed", "email": email, "previous_source_row": previous[email]["source_row"]})
        for email, fields in changed:
            writer.writerow({
                "change_type": "changed",
                "email": email,
                "previous_source_row": previous[email]["source_row"],
                "current_source_row": current[email]["source_row"],
                "changed_fields": ";".join(fields),
            })


if __name__ == "__main__":
    main()
